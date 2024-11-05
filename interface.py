import os.path
import tkinter
from tkinter import *
import customtkinter
import sys
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
import openpyxl
import threading
from main import generate_report
import re, zipfile

class MainWindow:
    def __init__(self):
        self.createWindow()

    def createWindow(self):


        def on_closing():
            if messagebox.askyesno("Salir","Seguro que quieres salir?"):
                self.app.destroy()
                print("Adios....")
                sys.exit()

        #customtkinter.set_appearance_mode("System")  # Modes: system (default), light, dark
        customtkinter.set_appearance_mode("Dark")
        #customtkinter.set_appearance_mode("Light")
        self.app = customtkinter.CTk()  # create CTk window like you do with the Tk window
        self.app.protocol("WM_DELETE_WINDOW",on_closing)
        self.app.title("Codigo SAP")
        self.app.state('normal')
        windowWidth = 0
        windowHeight = 0
        positionRight = int((self.app.winfo_screenwidth() / 2) - ((windowWidth + 12) / 2))
        positionDown = int((self.app.winfo_screenheight() / 2) - ((windowHeight + 50) / 2))
        self.app.geometry('{}x{}+{}+{}'.format(windowWidth, windowHeight, positionRight, positionDown))
        self.app.minsize(650,250)
        self.filePanel = customtkinter.CTkFrame(master=self.app)
        self.filePanel.pack(padx=10, pady=(20, 10), fill="both",
                       expand=False)  # .grid(row=0, column=0, padx=10, pady=(20, 10))
        self.filePanel.rowconfigure(0, weight=1)
        self.filePanel.rowconfigure(1, weight=1)
        self.filePanel.columnconfigure(0, weight=1)
        self.filePanel.columnconfigure(1, weight=1)
        #self.actionPanel = customtkinter.CTkFrame(master=self.app)
        #self.actionPanel.pack(padx=10, pady=(20, 10), fill="both", expand=False)
        #self.actionPanel.rowconfigure(0, weight=1)
        #self.actionPanel.columnconfigure(0, weight=1)
        #     image = tk.PhotoImage(file="imagen.png")
        # # Insertarla en una etiqueta.
        # label = ttk.Label(image=image)
        # label.pack()
                
        #label = customtkinter.CTkLabel(master=self.app, text="DOC Digital Transformation Team", font=("Roboto", 24))
        #label.pack()
        #card_report
        self.filePathcard_report = StringVar(value= "Seleccione el archivo Anexo")
        self.entrycard_report = customtkinter.CTkEntry(master=self.filePanel,
                                            width=300,
                                            placeholder_text="Seleccione el archivo Anexo",
                                            textvariable=self.filePathcard_report,
                                            justify='center')
        self.entrycard_report.grid(row=0, column=0, pady=20, padx=20, ipadx=150, sticky="ew")
        #self.entrycard_report.configure(fg='blue')
        #.tag_config("1", foreground="red")
        #self.entrycard_report.tag_config("1", foreground="blue")

        #sfp_report
        self.filePathsfp_report = StringVar(value= "Seleccione el archivo Base")
        self.entrysfp_report = customtkinter.CTkEntry(master=self.filePanel,
                                            width=300,
                                            placeholder_text="Seleccione el archivo Base",
                                            textvariable=self.filePathsfp_report,
                                            justify='center')

        self.entrysfp_report.grid(row=1, column=0, pady=20, padx=20, ipadx=150, sticky='ew')

       
        self.entrycard_report.bind("<1>", lambda name: self.openPath("Anexo"))
        self.entrysfp_report.bind("<1>", lambda name: self.openPath("Base"))
        
        
        self.button_5 = customtkinter.CTkButton(master=self.filePanel,
                                                text="Inicar Proceso",
                                                border_width=2,  # <- custom border_width
                                                command=self.startProcess)
        
        
        self.button_5.grid(row=4, column=0, columnspan=2, pady=20, padx=20, sticky="ew")
        self.app.mainloop()

    def openPath(self, file):
        try:
            if file == 'Anexo':
                self.entrycard_report.delete(0, 'end')
                self.entrycard_report.insert(0, "Selecciona el archivo Anexo")
                self.Pathcard_report = askopenfilename()
                if self.Pathcard_report == '':
                    self.entrycard_report.delete(0, 'end')
                    self.entrycard_report.insert(0, "Selecciona el archivo Anexo")
                else:
                    self.filePathcard_report.set(self.Pathcard_report.split('/')[-1])
                    #wb = openpyxl.load_workbook(filename=self.pathPEP, read_only=True, keep_links=False)
                    wb = self.getSheetNames(self.Pathcard_report)
                    self.hojaPathcard = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    self.hojaPathcard.grid(row=0, column=1, columnspan=1, pady=10, padx=20, ipadx=50, sticky="we")

                    #wb = openpyxl.load_workbook(filename=self.pathPEP, read_only=True, keep_links=False)
                    #wb = self.getSheetNames(self.pathPEP)
                    #self.hojaPEP = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    #self.hojaPEP.grid(row=0, column=1, columnspan=1, pady=10, padx=20, ipadx=50, sticky="we")
            if file == 'Base':
                self.entrysfp_report.delete(0, 'end')
                self.entrysfp_report.insert(0, "Selecciona el archivo Base")
                self.Pathsfp_report = askopenfilename()
                if self.Pathsfp_report == '':
                    self.entrysfp_report.delete(0, 'end')
                    self.entrysfp_report.insert(0, "Selecciona el archivo Base")
                else:
                    self.filePathsfp_report.set(self.Pathsfp_report.split('/')[-1])
                    wb = self.getSheetNames(self.Pathsfp_report)
                    self.hojaPathsfp = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    self.hojaPathsfp.grid(row=1, column=1, columnspan=1, pady=10, padx=20, ipadx=50, sticky="we")
                    #wb = openpyxl.load_workbook(filename=self.pathPEP, read_only=True, keep_links=False)
                    #wb = self.getSheetNames(self.pathPEP)
                    #self.hojaPEP = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    #self.hojaPEP.grid(row=0, column=1, columnspan=1, pady=10, padx=20, ipadx=50, sticky="we")

        except Exception as e:
            print(e)

    def getSheetNames(self, file_path):
        sheets = []
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            xml = zip_ref.read("xl/workbook.xml").decode("utf-8")
            for s_tag in re.findall("<sheet [^>]*", xml):
                sheets.append(re.search('name="[^"]*', s_tag).group(0)[6:])
        return sheets


    def startProcess(self):
        try:                                                #card_report,sfp_report,atp_Inventario
            threading.Thread(target=generate_report, args=(self.Pathcard_report,self.hojaPathcard.get(), self.Pathsfp_report,self.hojaPathsfp.get())).start()
            #threading.Thread(target=generate_report, args=(self.card_report, self.sfp_report, self.atp_Inventario)).start()
        except Exception as e:
            print(e)
            messagebox.showerror("Error", "Ha ocurrido un error inesperado. Por favor, intenta de nuevo o contacta a jose")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    a = MainWindow()

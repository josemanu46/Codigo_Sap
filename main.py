import win32com.client as client
import sys
import os
import time
import pdb
from tkinter import messagebox
import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
import logging
from styleframe import StyleFrame
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill
import linecache
from datetime import date, datetime
import pywintypes
import win32com.client as win32
import warnings
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from xlsxwriter.workbook import Workbook
from openpyxl.styles import PatternFill
export_path = os.path.join(current_dir, 'output')

def generate_report(materiales,sheet_materiales,stockSAP,sheet_stock):
    print('Start Generate report...')
    #pdb.set_trace()
    
    stock_SAP = pd.read_excel(stockSAP, sheet_name=sheet_stock,header=0, dtype=str)
    materiales_formato = pd.read_excel(materiales, sheet_name= sheet_materiales, header=13,dtype=str)
    materiales_formato.columns.str.match("Unnamed")
    
    df_materiales = materiales_formato.loc[:,~materiales_formato.columns.str.match("Unnamed")]

    # Realizar una fusión (merge) izquierda para combinar ambos DataFrames en función del "no serie"
    #merged_df = df_materiales.merge(stock_SAP, on='SERIE', how='left', suffixes=('', '_stock_SAP'))
    # Crear un DataFrame temporal para realizar la comparación
    #temp_df = excel2.copy()

    # eliminar nan, 
    df_stock_SAP = stock_SAP.dropna(subset=['SERIE'])
    
    unique_values = df_materiales['N° SERIE'].drop_duplicates().reset_index(drop=True)
    list = unique_values.tolist()
    common_values1 = df_stock_SAP['SERIE'][df_stock_SAP['SERIE'].isin(unique_values)]
    #.isin(unique_values)

    unique_values_df = common_values1.drop_duplicates().reset_index(drop=True)

    alexis =  df_stock_SAP['CODSAP'].to_list()
    list = unique_values_df.to_list()

    #Realizar una com de "no serie" en ambos df
    merged_df = df_materiales.merge(stock_SAP, left_on='N° SERIE', right_on='SERIE', how='left', suffixes=('', '_excel1'))

    # Definir condiciones para la coloración
    condicion__coincide = (merged_df['CODIGO SAP'].isna() | (merged_df['CODIGO SAP'] == merged_df['CODSAP']))
    print(condicion__coincide)
    # Colorea los valores comunes en HTML
 
    # def color_common_values(val):
    #     if val in common_values.tolist():
    #         return f"background-color: yellow"
    #     else:S
    #         return ""
    #condicion_codigo_nulo = merged_df['CODIGO SAP'].isna()
    # Colorea los valores comunes en HTML
    # Aplicar el formato de color en función de las condiciones

    def color_rows(val):
        if condicion__coincide[val.name]:
            return f"T"
        return "F"


    temp_df = df_materiales.copy()

    df_materiales['N° SERIE'] = df_materiales.apply(color_rows, axis=1)

    #pdb.set_trace()

    temp_df.rename(columns = {'N° SERIE':'TEST'}, inplace = True)

    df = pd.concat([temp_df['TEST'], df_materiales], axis=1)
    print(df)
    df = df.reindex(columns=['ITEM' ,'SOT' ,'CODIGO SAP',  'DESCRIPCION',  'UMB' ,'CANT. UTILIZADA', 'N° SERIE','TEST','PEDIDO', 'AREA CLARO', 'N° GUIA CLARO', 'OBSERVACIONE'])

    df.rename(columns = {'N° SERIE':'TEST1'}, inplace = True)
    df.rename(columns = {'TEST':'N° SERIE'}, inplace = True)

    df['N° SERIE']= df['N° SERIE'].fillna('N/A')
    #df.rename(columns = {'N° SERIE':'TEST'}, inplace = True)
    # df.insert(1, 'X', 0)
    #pdb.set_trace()
    # df
    # df['X'] = [9, 5, 1, 7, 3]

    # df
    #df['newcolumn'] = df['N° SERIE'] +df['TEST'] 

    #df['newcolumn'] = df['newcolumn'].astype(str)

    #print(df)
    
    #pdb.set_trace()
    
    #COLOREA MMDAS
    export_file = os.path.join(export_path, f'exported_test.xlsx')

    writer = pd.ExcelWriter(export_file, engine='openpyxl')


    #print(styled_df1)
    
    #sf.to_csv(export_file, index=False)
    #excel_writer = StyleFrame.ExcelWriter(export_file)
    #sf = StyleFrame(df_combined)
    #sf.to_frame()
    #df.to_excel(excel_writer=export_file,index=False)
    print('export')
    # Descombinar todas las celdas en la hoja
    #df_frame_full = pd.read_excel(export_file, sheet_name= 0, header=0)

    df.to_excel(writer, sheet_name='Sheet1', index=False)

    book = writer.book
    sheet = writer.sheets['Sheet1']

    blue_fill = PatternFill(start_color='0FEA30', end_color='0FEA30', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFC300', end_color='FFC300', fill_type='solid')
    red_fill = PatternFill(start_color='EE4B2B', end_color='EE4B2B', fill_type='solid')

    #Aplicar estilos
    for idx, row in df.iterrows():
        if row['TEST1'] == 'T' and row['N° SERIE'] in list:
            cell = sheet.cell(row=idx + 2, column=8)  # +2  excel comienza desde la fila 2
            cell.fill = blue_fill
        elif row['TEST1'] == 'F' and row['N° SERIE'] in list:
            cell = sheet.cell(row=idx + 2, column=8)
            cell.fill = yellow_fill
        elif row['TEST1'] == 'F' and row['N° SERIE'] != list:
            cell = sheet.cell(row=idx + 2, column=8)
            cell.fill = red_fill        

    #eliminamos la col TEST
    sheet.delete_cols(7)

    

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    writer.save()
    messagebox.showinfo("Codigo SAP", "Archivo creado exitosamente!!")
    abrir_ubicacion(export_path)
    email_outlook()
    #messagebox.showinfo("Codigo SAP", "Archivo creado exitosamente!!")
    



def abrir_ubicacion(template):
    if template: 
        respuesta = messagebox.askyesno("Confirmar", "¿Desea abrir la carpeta?")
        if respuesta:
            try:
                os.startfile(template)
            except OSError as e:
                messagebox.showerror("Error", f"No se pudo abrir la ubicación de la carpeta:\n{str(e)}")
    else:
        messagebox.showinfo("Información", "No se ha seleccionado una ubicación de carpeta.")




def email_outlook():
    
    outlook = client.GetActiveObject('Outlook.Application')
    #print('Outlook abierto...')
    # except (pywintypes.com_error , AttributeError):
    #     outlook = None
    #     print('Outlook cerrado...')
        #self.PrintException()
    #cierra Outlook si hay una instancia activa
    # try:
    #     if outlook:
    #         outlook.Quit()
    # except pywintypes.com_error:
    #     pass

    try:
        outlook = client.Dispatch('Outlook.Application')
    except pywintypes.com_error:
        outlook = None
    if outlook:    
        try:
            #outlook = client.Dispatch('Outlook.Application')
            mail = outlook.CreateItem(0)  # Crear un nuevo correo electrónico
            # Obtener fecha actual
            fecha_actual = date(datetime.now().year, datetime.now().month, datetime.now().day).strftime("%d-%m-%Y")
            #html_Body = 
            # Configurar los campos del correo electrónico
            #mail.Subject = f"Daily report MRs in 'applied' status pending to approve {fecha_actual}"

            warnings.filterwarnings("ignore", category=UserWarning)
            #pdb.set_trace()
            # Convertir las listas de destinatarios en cadenas separadas por comas
            #to_emails = ";".join(recipients)
            #cc_emails = ";".join(cc_recipients)
            #mail.To = "; ".join(to_emails[0])
            #mail.Cc = "; ".join(cc_emails[0])
            #mail.To = to_emails
            #mail.CC = cc_emails
            mail.Subject = f"Code SAP Tool"
            mail.Body = f"Tool utilizada:  {fecha_actual}"

            mail.Send()
            print("Correo electrónico enviado exitosamente")
        finally:
            mail = None
    else:
        print("no se pudo crear ni obtener outlook")








#generate_report(stockSAP,materiales)